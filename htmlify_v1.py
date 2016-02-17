#Kenneth Ng (C) 2016
#Version 0.1
#Script to convert single column of string data in excel to bulleted list

import openpyxl
from openpyxl import load_workbook
import os


descriptions = []
html_list = []


# ask user for number of rows in document and document name
print('*******************************************************\n')
print('      Welcome to HTMLify created by Kenneth Ng.\n')
print('*******************************************************\n\n')
print('Make sure data to be transformed is in column A of .xlsx file.\nEnsure this file is saved to same folder as htmlify.exe\n')
user_input = input('1. Enter the number of rows in document: ')
num_rows = 'A1:A'+user_input
file_input = input('2. Specify the filename (e.g. document.xlsx): ')
wb = load_workbook(os.getcwd()+'\\'+file_input)

	



# method to import strings, split text into list items using '\n' and add <ul>, <li> and </li> to items
# join list items back into string

def htmlify(input_string):
	item_list = input_string.split('\n')
	html_list = ['<li>'+x+'</li>' for x in item_list]
	html_list.insert(0, '<ul>')
	html_list.append('</ul>')
	return '\n'.join(html_list)
	


ws = wb.active
descriptions = [str(cell.value) for row in ws.iter_rows(num_rows) for cell in row]
html_descriptions = [htmlify(x) for x in descriptions]



#create new worksheet 'processed' and write list data into column A1
ws2 = wb.create_sheet(title='processed')

count = 0
for row in ws2.iter_rows(num_rows):
	for cell in row:
		cell.value = html_descriptions[count]
		count = count + 1

print('\nReading.. '+file_input)
print('Transforming data')
print('Writing to worksheet \"processed\" with html-ified lists')
print('Saving to ... processed--'+file_input)

	
# save xlsx data to new workbook
wb.save(os.getcwd()+'\\processed--'+file_input)
print('Complete!\n')

leave = input('Press ENTER key to continue.')
